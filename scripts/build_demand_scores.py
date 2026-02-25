#!/usr/bin/env python3
"""
Build demand scores per statistical sector.

Combines:
1. Age demographics (municipality level) — % population aged 25-54 (peak e-commerce)
2. Fiscal income (municipality level) — average net taxable income per capita

Formula: demand_score = population × ecommerce_age_ratio × income_index

Output: Enriches data/centroids.json with 'demand' field per sector.
Also outputs data/demand_scores.json for reference.
"""

import json
import csv
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

DATA_DIR = Path("/Users/henry/Desktop/bbox-coverage-tool/data")

# E-commerce peak age range
ECOM_AGE_MIN = 25
ECOM_AGE_MAX = 54


def load_age_ratios():
    """
    Load age demographics from StatBel TF_SOC_POP_STRUCT_2025.txt.
    Returns dict: {municipality_nis: ecommerce_age_ratio}
    """
    print("Loading age demographics...")
    age_file = DATA_DIR / "statbel_age_sex" / "TF_SOC_POP_STRUCT_2025.txt"

    # Aggregate population by municipality and age bucket
    mun_total = defaultdict(int)
    mun_ecom = defaultdict(int)

    with open(age_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            mun = row["CD_REFNIS"]
            age = int(row["CD_AGE"])
            pop = int(row["MS_POPULATION"])

            mun_total[mun] += pop
            if ECOM_AGE_MIN <= age <= ECOM_AGE_MAX:
                mun_ecom[mun] += pop

    # Compute ratios
    ratios = {}
    for mun in mun_total:
        total = mun_total[mun]
        if total > 0:
            ratios[mun] = mun_ecom[mun] / total
        else:
            ratios[mun] = 0.0

    avg_ratio = sum(ratios.values()) / len(ratios) if ratios else 0.0
    print(f"  {len(ratios)} municipalities loaded")
    print(f"  Average e-commerce age ratio (25-54): {avg_ratio:.3f}")

    return ratios, avg_ratio


def load_income_per_capita():
    """
    Load fiscal income from StatBel TF_PSNL_INC_TAX_MUNTY.xlsx.
    Uses most recent year available.
    Returns dict: {municipality_nis: avg_income_per_capita}
    """
    print("Loading fiscal income data...")
    income_file = DATA_DIR / "statbel_fiscal_income.xlsx"

    wb = openpyxl.load_workbook(income_file, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Read all rows, keep only the most recent year per municipality
    header = None
    mun_data = {}  # {mun_nis: {year, total_income, residents}}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue

        year = row[0]
        mun_nis = str(row[1])
        total_income = row[4]  # MS_TOT_NET_TAXABLE_INC
        residents = row[29]    # MS_TOT_RESIDENTS

        if year is None or total_income is None or residents is None:
            continue

        # Keep most recent year
        if mun_nis not in mun_data or year > mun_data[mun_nis]["year"]:
            mun_data[mun_nis] = {
                "year": year,
                "total_income": float(total_income),
                "residents": int(residents),
            }

    wb.close()

    # Compute income per capita
    income_pc = {}
    for mun, d in mun_data.items():
        if d["residents"] > 0:
            income_pc[mun] = d["total_income"] / d["residents"]

    # National average for normalisation
    total_inc = sum(d["total_income"] for d in mun_data.values())
    total_res = sum(d["residents"] for d in mun_data.values())
    national_avg = total_inc / total_res if total_res > 0 else 1.0

    most_recent_year = max(d["year"] for d in mun_data.values())
    print(f"  {len(income_pc)} municipalities loaded (most recent year: {most_recent_year})")
    print(f"  National average income per capita: EUR {national_avg:,.0f}")

    return income_pc, national_avg


def extract_municipality_nis(sector_code):
    """Extract 5-digit municipality NIS code from sector code."""
    return sector_code[:5]


def build_district_fallbacks(data_dict, avg_fallback):
    """
    Build district-level averages as fallback for municipalities missing
    from the data (due to 2025 municipal mergers).
    District = first 2 digits of NIS code.
    """
    district_vals = defaultdict(list)
    for mun, val in data_dict.items():
        district = mun[:2]
        district_vals[district].append(val)

    district_avg = {}
    for d, vals in district_vals.items():
        district_avg[d] = sum(vals) / len(vals)

    return district_avg


def main():
    # Load source data
    age_ratios, avg_age_ratio = load_age_ratios()
    income_pc, national_avg_income = load_income_per_capita()

    # Build district-level fallbacks for merged municipalities
    district_age = build_district_fallbacks(age_ratios, avg_age_ratio)
    income_idx_dict = {m: v / national_avg_income for m, v in income_pc.items()}
    district_income = build_district_fallbacks(income_idx_dict, 1.0)

    # Load centroids
    print("\nLoading centroids...")
    with open(DATA_DIR / "centroids.json", "r") as f:
        centroids = json.load(f)

    print(f"  {len(centroids)} sectors")

    # Compute demand scores
    print("\nComputing demand scores...")
    scores = []
    missing_age = 0
    missing_income = 0
    fallback_age = 0
    fallback_income = 0

    for c in centroids:
        mun_nis = extract_municipality_nis(c["sc"])
        district = mun_nis[:2]
        pop = c["pop"]

        # Age ratio: try municipality -> district -> national average
        age_ratio = age_ratios.get(mun_nis)
        if age_ratio is None:
            age_ratio = district_age.get(district)
            if age_ratio is not None:
                fallback_age += 1
            else:
                age_ratio = avg_age_ratio
                missing_age += 1

        # Income index: try municipality -> district -> 1.0
        ipc = income_pc.get(mun_nis)
        if ipc is None:
            income_index = district_income.get(district)
            if income_index is not None:
                fallback_income += 1
            else:
                income_index = 1.0
                missing_income += 1
        else:
            income_index = ipc / national_avg_income

        # Demand score
        demand = pop * age_ratio * income_index
        scores.append(round(demand, 1))

        # Add to centroid
        c["demand"] = round(demand, 1)
        c["ageRatio"] = round(age_ratio, 4)
        c["incomeIdx"] = round(income_index, 3)

    if fallback_age > 0:
        print(f"  {fallback_age} sectors used district-level age fallback (municipal mergers)")
    if missing_age > 0:
        print(f"  WARNING: {missing_age} sectors missing age data entirely (used national average)")
    if fallback_income > 0:
        print(f"  {fallback_income} sectors used district-level income fallback (municipal mergers)")
    if missing_income > 0:
        print(f"  WARNING: {missing_income} sectors missing income data entirely (used index 1.0)")

    # Stats
    nonzero = [s for s in scores if s > 0]
    print(f"\n  Demand score stats:")
    print(f"    Non-zero sectors: {len(nonzero)} / {len(scores)}")
    print(f"    Min: {min(nonzero):.1f}")
    print(f"    Max: {max(nonzero):.1f}")
    print(f"    Mean: {sum(nonzero)/len(nonzero):.1f}")
    print(f"    Median: {sorted(nonzero)[len(nonzero)//2]:.1f}")
    total_demand = sum(scores)
    print(f"    Total demand: {total_demand:,.0f}")

    # Save enriched centroids
    print("\nSaving enriched centroids.json...")
    with open(DATA_DIR / "centroids.json", "w") as f:
        json.dump(centroids, f, separators=(",", ":"))

    # Save standalone demand reference file
    demand_ref = []
    for c in centroids:
        demand_ref.append({
            "sc": c["sc"],
            "mun": extract_municipality_nis(c["sc"]),
            "pop": c["pop"],
            "ageRatio": c["ageRatio"],
            "incomeIdx": c["incomeIdx"],
            "demand": c["demand"],
        })

    with open(DATA_DIR / "demand_scores.json", "w") as f:
        json.dump(demand_ref, f, separators=(",", ":"))

    ref_size = (DATA_DIR / "demand_scores.json").stat().st_size / 1024
    print(f"  demand_scores.json: {ref_size:.0f} KB")
    print("\nDone!")


if __name__ == "__main__":
    main()
